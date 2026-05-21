"""
Generate an Excel file documenting all database entity schemas,
grouped into three logical categories.
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------------------
# Schema definitions
# Each column dict: name, type, pk, index, nullable
# BaseEntity fields shared by all tables: id, created_at, updated_at
# ---------------------------------------------------------------------------

BASE_COLS = [
    {"name": "id", "type": "UUID", "pk": "yes", "index": "yes", "nullable": "no"},
    {
        "name": "created_at",
        "type": "TIMESTAMP",
        "pk": "no",
        "index": "yes",
        "nullable": "no",
    },
    {
        "name": "updated_at",
        "type": "TIMESTAMP",
        "pk": "no",
        "index": "no",
        "nullable": "no",
    },
]

GROUPS = [
    {
        "label": "Group 1: System Table",
        "color": "4472C4",  # blue
        "tables": [
            {
                "name": "users",
                "columns": BASE_COLS
                + [
                    {
                        "name": "name",
                        "type": "VARCHAR(255)",
                        "pk": "no",
                        "index": "no",
                        "nullable": "no",
                    },
                    {
                        "name": "email",
                        "type": "VARCHAR(255)",
                        "pk": "no",
                        "index": "yes",
                        "nullable": "no",
                    },
                    {
                        "name": "password_hash",
                        "type": "VARCHAR(255)",
                        "pk": "no",
                        "index": "no",
                        "nullable": "yes",
                    },
                    {
                        "name": "role",
                        "type": "VARCHAR(50)",
                        "pk": "no",
                        "index": "no",
                        "nullable": "no",
                    },
                    {
                        "name": "is_email_verified",
                        "type": "BOOLEAN",
                        "pk": "no",
                        "index": "no",
                        "nullable": "no",
                    },
                    {
                        "name": "email_verification_sent_at",
                        "type": "TIMESTAMP",
                        "pk": "no",
                        "index": "no",
                        "nullable": "yes",
                    },
                    {
                        "name": "last_login_at",
                        "type": "TIMESTAMP",
                        "pk": "no",
                        "index": "no",
                        "nullable": "yes",
                    },
                ],
            },
            {
                "name": "audit_logs",
                "columns": BASE_COLS
                + [
                    {
                        "name": "actor_type",
                        "type": "VARCHAR(50)",
                        "pk": "no",
                        "index": "no",
                        "nullable": "no",
                    },
                    {
                        "name": "actor_id",
                        "type": "UUID",
                        "pk": "no",
                        "index": "no",
                        "nullable": "yes",
                    },
                    {
                        "name": "entity_type",
                        "type": "VARCHAR(100)",
                        "pk": "no",
                        "index": "no",
                        "nullable": "no",
                    },
                    {
                        "name": "entity_id",
                        "type": "UUID",
                        "pk": "no",
                        "index": "no",
                        "nullable": "yes",
                    },
                    {
                        "name": "action_type",
                        "type": "VARCHAR(50)",
                        "pk": "no",
                        "index": "no",
                        "nullable": "no",
                    },
                    {
                        "name": "before_data",
                        "type": "TEXT",
                        "pk": "no",
                        "index": "no",
                        "nullable": "yes",
                    },
                    {
                        "name": "after_data",
                        "type": "TEXT",
                        "pk": "no",
                        "index": "no",
                        "nullable": "yes",
                    },
                    {
                        "name": "request_ip",
                        "type": "VARCHAR(100)",
                        "pk": "no",
                        "index": "no",
                        "nullable": "yes",
                    },
                    {
                        "name": "client",
                        "type": "VARCHAR(255)",
                        "pk": "no",
                        "index": "no",
                        "nullable": "yes",
                    },
                ],
            },
            {
                "name": "file_metadatas",
                "columns": BASE_COLS
                + [
                    {
                        "name": "name",
                        "type": "VARCHAR(255)",
                        "pk": "no",
                        "index": "no",
                        "nullable": "no",
                    },
                    {
                        "name": "path",
                        "type": "VARCHAR(1024)",
                        "pk": "no",
                        "index": "no",
                        "nullable": "no",
                    },
                    {
                        "name": "size",
                        "type": "BIGINT",
                        "pk": "no",
                        "index": "no",
                        "nullable": "no",
                    },
                    {
                        "name": "mime_type",
                        "type": "VARCHAR(255)",
                        "pk": "no",
                        "index": "no",
                        "nullable": "yes",
                    },
                    {
                        "name": "object_key",
                        "type": "VARCHAR(255)",
                        "pk": "no",
                        "index": "no",
                        "nullable": "yes",
                    },
                ],
            },
        ],
    },
    {
        "label": "Group 2: Document ETL Table",
        "color": "70AD47",  # green
        "tables": [
            {
                "name": "documents",
                "columns": BASE_COLS
                + [
                    {
                        "name": "name",
                        "type": "VARCHAR(255)",
                        "pk": "no",
                        "index": "no",
                        "nullable": "no",
                    },
                    {
                        "name": "file_id",
                        "type": "VARCHAR(255)",
                        "pk": "no",
                        "index": "no",
                        "nullable": "no",
                    },
                    {
                        "name": "status",
                        "type": "VARCHAR(50)",
                        "pk": "no",
                        "index": "no",
                        "nullable": "no",
                    },
                    {
                        "name": "progress",
                        "type": "FLOAT",
                        "pk": "no",
                        "index": "no",
                        "nullable": "no",
                    },
                    {
                        "name": "metadata",
                        "type": "JSONB",
                        "pk": "no",
                        "index": "no",
                        "nullable": "yes",
                    },
                    {
                        "name": "uploaded_by_id",
                        "type": "UUID (FK → users.id)",
                        "pk": "no",
                        "index": "yes",
                        "nullable": "yes",
                    },
                ],
            },
            {
                "name": "pages",
                "columns": BASE_COLS
                + [
                    {
                        "name": "document_id",
                        "type": "UUID (FK → documents.id)",
                        "pk": "no",
                        "index": "no",
                        "nullable": "no",
                    },
                    {
                        "name": "page_number",
                        "type": "INTEGER",
                        "pk": "no",
                        "index": "no",
                        "nullable": "no",
                    },
                    {
                        "name": "content",
                        "type": "TEXT",
                        "pk": "no",
                        "index": "no",
                        "nullable": "yes",
                    },
                    {
                        "name": "validated_content",
                        "type": "TEXT",
                        "pk": "no",
                        "index": "no",
                        "nullable": "yes",
                    },
                    {
                        "name": "overlap_content",
                        "type": "TEXT",
                        "pk": "no",
                        "index": "no",
                        "nullable": "yes",
                    },
                    {
                        "name": "page_image_id",
                        "type": "UUID",
                        "pk": "no",
                        "index": "no",
                        "nullable": "yes",
                    },
                    {
                        "name": "image_list",
                        "type": "JSONB",
                        "pk": "no",
                        "index": "no",
                        "nullable": "yes",
                    },
                ],
            },
            {
                "name": "questions",
                "columns": BASE_COLS
                + [
                    {
                        "name": "page_id",
                        "type": "UUID (FK → pages.id)",
                        "pk": "no",
                        "index": "no",
                        "nullable": "yes",
                    },
                    {
                        "name": "parent_question_id",
                        "type": "UUID (FK → questions.id)",
                        "pk": "no",
                        "index": "no",
                        "nullable": "yes",
                    },
                    {
                        "name": "questions_group_id",
                        "type": "UUID (FK → questions_groups.id)",
                        "pk": "no",
                        "index": "no",
                        "nullable": "yes",
                    },
                    {
                        "name": "question_text",
                        "type": "TEXT",
                        "pk": "no",
                        "index": "no",
                        "nullable": "no",
                    },
                    {
                        "name": "question_type",
                        "type": "VARCHAR(50)",
                        "pk": "no",
                        "index": "no",
                        "nullable": "no",
                    },
                    {
                        "name": "difficulty",
                        "type": "VARCHAR(50)",
                        "pk": "no",
                        "index": "no",
                        "nullable": "yes",
                    },
                    {
                        "name": "subject",
                        "type": "VARCHAR(255)",
                        "pk": "no",
                        "index": "yes",
                        "nullable": "yes",
                    },
                    {
                        "name": "topic",
                        "type": "VARCHAR(255)",
                        "pk": "no",
                        "index": "yes",
                        "nullable": "yes",
                    },
                    {
                        "name": "image_list",
                        "type": "JSONB",
                        "pk": "no",
                        "index": "no",
                        "nullable": "yes",
                    },
                    {
                        "name": "sub_question_order",
                        "type": "INTEGER",
                        "pk": "no",
                        "index": "no",
                        "nullable": "yes",
                    },
                    {
                        "name": "variant_existence_count",
                        "type": "BIGINT",
                        "pk": "no",
                        "index": "no",
                        "nullable": "no",
                    },
                    {
                        "name": "vector_embedding",
                        "type": "VECTOR(768)",
                        "pk": "no",
                        "index": "no",
                        "nullable": "yes",
                    },
                    {
                        "name": "status",
                        "type": "SMALLINT",
                        "pk": "no",
                        "index": "no",
                        "nullable": "no",
                    },
                ],
            },
            {
                "name": "answers",
                "columns": BASE_COLS
                + [
                    {
                        "name": "question_id",
                        "type": "UUID (FK → questions.id)",
                        "pk": "no",
                        "index": "no",
                        "nullable": "no",
                    },
                    {
                        "name": "value",
                        "type": "TEXT",
                        "pk": "no",
                        "index": "no",
                        "nullable": "no",
                    },
                    {
                        "name": "is_correct",
                        "type": "BOOLEAN",
                        "pk": "no",
                        "index": "no",
                        "nullable": "no",
                    },
                    {
                        "name": "explaination",
                        "type": "TEXT",
                        "pk": "no",
                        "index": "no",
                        "nullable": "yes",
                    },
                ],
            },
            {
                "name": "questions_groups",
                "columns": BASE_COLS
                + [
                    {
                        "name": "subject",
                        "type": "VARCHAR(255)",
                        "pk": "no",
                        "index": "no",
                        "nullable": "no",
                    },
                    {
                        "name": "topic",
                        "type": "VARCHAR(255)",
                        "pk": "no",
                        "index": "no",
                        "nullable": "no",
                    },
                    {
                        "name": "difficulty",
                        "type": "VARCHAR(50)",
                        "pk": "no",
                        "index": "no",
                        "nullable": "no",
                    },
                    {
                        "name": "existence_count",
                        "type": "BIGINT",
                        "pk": "no",
                        "index": "no",
                        "nullable": "no",
                    },
                    {
                        "name": "vector_embedding",
                        "type": "VECTOR(768)",
                        "pk": "no",
                        "index": "no",
                        "nullable": "yes",
                    },
                    {
                        "name": "from_user_id",
                        "type": "UUID (FK → users.id)",
                        "pk": "no",
                        "index": "yes",
                        "nullable": "yes",
                    },
                ],
            },
            {
                "name": "topics",
                "columns": BASE_COLS
                + [
                    {
                        "name": "name",
                        "type": "VARCHAR(255)",
                        "pk": "no",
                        "index": "no",
                        "nullable": "no",
                    },
                    {
                        "name": "name_vi",
                        "type": "VARCHAR(255)",
                        "pk": "no",
                        "index": "no",
                        "nullable": "yes",
                    },
                    {
                        "name": "subject_code",
                        "type": "VARCHAR(50)",
                        "pk": "no",
                        "index": "yes",
                        "nullable": "yes",
                    },
                    {
                        "name": "code",
                        "type": "VARCHAR(50)",
                        "pk": "no",
                        "index": "yes",
                        "nullable": "no",
                    },
                    {
                        "name": "description",
                        "type": "TEXT",
                        "pk": "no",
                        "index": "no",
                        "nullable": "yes",
                    },
                ],
            },
            {
                "name": "subjects",
                "columns": BASE_COLS
                + [
                    {
                        "name": "name",
                        "type": "VARCHAR(255)",
                        "pk": "no",
                        "index": "no",
                        "nullable": "no",
                    },
                    {
                        "name": "name_vi",
                        "type": "VARCHAR(255)",
                        "pk": "no",
                        "index": "no",
                        "nullable": "yes",
                    },
                    {
                        "name": "code",
                        "type": "VARCHAR(50)",
                        "pk": "no",
                        "index": "yes",
                        "nullable": "no",
                    },
                    {
                        "name": "description",
                        "type": "TEXT",
                        "pk": "no",
                        "index": "no",
                        "nullable": "yes",
                    },
                ],
            },
        ],
    },
    {
        "label": "Group 3: Exam Generation + Exam Taking",
        "color": "ED7D31",  # orange
        "tables": [
            {
                "name": "attempt_token_mappings",
                "columns": BASE_COLS
                + [
                    {
                        "name": "token_hash",
                        "type": "VARCHAR(255)",
                        "pk": "no",
                        "index": "yes",
                        "nullable": "no",
                    },
                    {
                        "name": "attempt_id",
                        "type": "VARCHAR(255)",
                        "pk": "no",
                        "index": "yes",
                        "nullable": "no",
                    },
                    {
                        "name": "expires_at",
                        "type": "TIMESTAMP",
                        "pk": "no",
                        "index": "yes",
                        "nullable": "no",
                    },
                    {
                        "name": "is_invalidated",
                        "type": "BOOLEAN",
                        "pk": "no",
                        "index": "no",
                        "nullable": "no",
                    },
                ],
            },
            {
                "name": "exam_templates",
                "columns": BASE_COLS
                + [
                    {
                        "name": "name",
                        "type": "VARCHAR(255)",
                        "pk": "no",
                        "index": "no",
                        "nullable": "no",
                    },
                    {
                        "name": "subject",
                        "type": "VARCHAR(50)",
                        "pk": "no",
                        "index": "no",
                        "nullable": "no",
                    },
                    {
                        "name": "generation_config",
                        "type": "TEXT",
                        "pk": "no",
                        "index": "no",
                        "nullable": "yes",
                    },
                    {
                        "name": "created_by_id",
                        "type": "UUID (FK → users.id)",
                        "pk": "no",
                        "index": "yes",
                        "nullable": "yes",
                    },
                ],
            },
            {
                "name": "exam_instances",
                "columns": BASE_COLS
                + [
                    {
                        "name": "exam_template_id",
                        "type": "UUID (FK → exam_templates.id)",
                        "pk": "no",
                        "index": "yes",
                        "nullable": "yes",
                    },
                    {
                        "name": "parent_exam_instance_id",
                        "type": "UUID (FK → exam_instances.id)",
                        "pk": "no",
                        "index": "no",
                        "nullable": "yes",
                    },
                    {
                        "name": "created_by_id",
                        "type": "UUID (FK → users.id)",
                        "pk": "no",
                        "index": "yes",
                        "nullable": "yes",
                    },
                    {
                        "name": "exported_file_id",
                        "type": "VARCHAR(255)",
                        "pk": "no",
                        "index": "no",
                        "nullable": "yes",
                    },
                    {
                        "name": "exam_test_code",
                        "type": "VARCHAR(255)",
                        "pk": "no",
                        "index": "yes",
                        "nullable": "no",
                    },
                    {
                        "name": "is_exported",
                        "type": "BOOLEAN",
                        "pk": "no",
                        "index": "no",
                        "nullable": "no",
                    },
                    {
                        "name": "is_base",
                        "type": "BOOLEAN",
                        "pk": "no",
                        "index": "no",
                        "nullable": "no",
                    },
                    {
                        "name": "status",
                        "type": "SMALLINT",
                        "pk": "no",
                        "index": "yes",
                        "nullable": "no",
                    },
                ],
            },
            {
                "name": "exam_test_sections",
                "columns": BASE_COLS
                + [
                    {
                        "name": "exam_instance_id",
                        "type": "UUID (FK → exam_instances.id)",
                        "pk": "no",
                        "index": "no",
                        "nullable": "no",
                    },
                    {
                        "name": "name",
                        "type": "VARCHAR(255)",
                        "pk": "no",
                        "index": "no",
                        "nullable": "no",
                    },
                    {
                        "name": "order_index",
                        "type": "INTEGER",
                        "pk": "no",
                        "index": "no",
                        "nullable": "no",
                    },
                ],
            },
            {
                "name": "user_test_attempts",
                "columns": BASE_COLS
                + [
                    {
                        "name": "user_id",
                        "type": "UUID (FK → users.id)",
                        "pk": "no",
                        "index": "no",
                        "nullable": "no",
                    },
                    {
                        "name": "exam_template_id",
                        "type": "VARCHAR(255)",
                        "pk": "no",
                        "index": "no",
                        "nullable": "no",
                    },
                    {
                        "name": "exam_instance_id",
                        "type": "UUID (FK → exam_instances.id)",
                        "pk": "no",
                        "index": "no",
                        "nullable": "yes",
                    },
                    {
                        "name": "score",
                        "type": "DECIMAL(5,2)",
                        "pk": "no",
                        "index": "no",
                        "nullable": "yes",
                    },
                    {
                        "name": "status",
                        "type": "SMALLINT",
                        "pk": "no",
                        "index": "yes",
                        "nullable": "no",
                    },
                    {
                        "name": "started_at",
                        "type": "TIMESTAMP",
                        "pk": "no",
                        "index": "no",
                        "nullable": "no",
                    },
                    {
                        "name": "submitted_at",
                        "type": "TIMESTAMP",
                        "pk": "no",
                        "index": "no",
                        "nullable": "yes",
                    },
                ],
            },
            {
                "name": "user_test_attempt_answers",
                "columns": BASE_COLS
                + [
                    {
                        "name": "attempt_id",
                        "type": "UUID (FK → user_test_attempts.id)",
                        "pk": "no",
                        "index": "no",
                        "nullable": "no",
                    },
                    {
                        "name": "question_id",
                        "type": "VARCHAR(255)",
                        "pk": "no",
                        "index": "no",
                        "nullable": "no",
                    },
                    {
                        "name": "selected_answer_id",
                        "type": "VARCHAR(255)",
                        "pk": "no",
                        "index": "no",
                        "nullable": "yes",
                    },
                    {
                        "name": "is_correct",
                        "type": "BOOLEAN",
                        "pk": "no",
                        "index": "no",
                        "nullable": "no",
                    },
                    {
                        "name": "time_spent",
                        "type": "BIGINT",
                        "pk": "no",
                        "index": "no",
                        "nullable": "no",
                    },
                ],
            },
        ],
    },
]

# ---------------------------------------------------------------------------
# Styling helpers
# ---------------------------------------------------------------------------

HEADER_ROW = ["name", "type", "nullable", "purpose"]
COL_WIDTHS = [30, 36, 10, 50]


def hex_fill(hex_color: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_color)


def thin_border() -> Border:
    side = Side(style="thin", color="BFBFBF")
    return Border(left=side, right=side, top=side, bottom=side)


def apply_col_widths(ws):
    for i, w in enumerate(COL_WIDTHS, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def write_group_header(ws, row: int, label: str, color: str) -> int:
    cell = ws.cell(row=row, column=1, value=label)
    cell.font = Font(bold=True, color="FFFFFF", size=12)
    cell.fill = hex_fill(color)
    cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.merge_cells(
        start_row=row, start_column=1, end_row=row, end_column=len(HEADER_ROW)
    )
    ws.row_dimensions[row].height = 22
    return row + 1


def write_table_header(ws, row: int, table_name: str, color: str) -> int:
    # Table name row
    name_cell = ws.cell(row=row, column=1, value=f"  {table_name}")
    name_cell.font = Font(bold=True, color="FFFFFF", size=11)
    name_cell.fill = hex_fill(
        color + "BF"[0:] if len(color) == 6 else color
    )  # same color, lighter feel via tint below
    # Use a slightly lighter shade by appending alpha — openpyxl uses aRGB so prefix FF
    light_fill = PatternFill("solid", fgColor="FF" + color)
    name_cell.fill = light_fill
    name_cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.merge_cells(
        start_row=row, start_column=1, end_row=row, end_column=len(HEADER_ROW)
    )
    ws.row_dimensions[row].height = 20
    row += 1

    # Column header row
    light_hex = _lighten(color)
    for col_idx, h in enumerate(HEADER_ROW, start=1):
        cell = ws.cell(row=row, column=col_idx, value=h)
        cell.font = Font(bold=True, size=10)
        cell.fill = hex_fill(light_hex)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border()
    ws.row_dimensions[row].height = 16
    return row + 1


def _lighten(hex_color: str) -> str:
    """Return a lighter (pastel) version of the hex color by blending toward white."""
    r = int(hex_color[0:2], 16)
    g = int(hex_color[2:4], 16)
    b = int(hex_color[4:6], 16)
    factor = 0.55
    r2 = int(r + (255 - r) * factor)
    g2 = int(g + (255 - g) * factor)
    b2 = int(b + (255 - b) * factor)
    return f"{r2:02X}{g2:02X}{b2:02X}"


def write_columns(ws, row: int, columns: list) -> int:
    for col_def in columns:
        values = [
            col_def["name"],
            col_def["type"],
            col_def["nullable"],
            "",  # purpose — left empty for manual entry
        ]
        for col_idx, val in enumerate(values, start=1):
            cell = ws.cell(row=row, column=col_idx, value=val)
            cell.font = Font(size=10)
            cell.alignment = Alignment(
                horizontal="left" if col_idx <= 2 else "center", vertical="center"
            )
            cell.border = thin_border()
            if col_def.get("pk") == "yes" and col_idx == 1:
                cell.font = Font(size=10, bold=True, color="7030A0")
        ws.row_dimensions[row].height = 15
        row += 1
    return row


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------


def generate(output_path: str = "db_schema.xlsx"):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "DB Schema"
    ws.freeze_panes = "A2"

    apply_col_widths(ws)

    row = 1
    for group in GROUPS:
        color = group["color"]
        row = write_group_header(ws, row, group["label"], color)

        for table in group["tables"]:
            row = write_table_header(ws, row, table["name"], color)
            row = write_columns(ws, row, table["columns"])
            row += 1  # blank row between tables

        row += 1  # extra blank row between groups

    wb.save(output_path)
    print(f"Saved: {output_path}")


if __name__ == "__main__":
    import sys

    path = sys.argv[1] if len(sys.argv) > 1 else "db_schema.xlsx"
    generate(path)
